import numpy as np
import math
import openpyxl
import csv
from datetime import datetime



##############################################



#Saving data into excel sheets every N measurements - saving 1 row at a time
class StoreData():

    def __init__(self,header,att):
        self.header = header
        self.att = att
        # decides if it will create a csv file or xls file depending on the parent class att super definition
        
        if self.att == 'csvf':
            #Creating a csv
            with open("data.csv" , "w") as csv_file_obj:
                csv_writer = csv.writer(csv_file_obj, delimiter = ",")
                csv_writer.writerow(['Current_time' , 'Minimum_Temperature_Confidence_Interval' , 
                         'Maximun_Temperature_Confidence_Interval',
                         'Mean_Temperature' , 'Median_Temparature' , 'Minimum_Pressure_Confidence_Interval' , 
                         'Maximun_Pressure_Confidence_Interval' ,'Mean_Pressure' , 'Median_Pressure' ,
                         'Minimum_Humidity_Confidence_Interval' , 'Maximun_Humidity_Confidence_Interval',
                         'Mean_Humidity' , 'Median_Humidity'])
                csv_file_obj.close()
                
                   
        elif self.att == 'xls':
            #We re creating or opening a new xlsx file with the folowing 2 com
            wb = openpyxl.Workbook() 
            wb.save("data.xlsx")
            #print(self.header) #header check

            #Creation of necessary sheets in case the file is freshly created
            #also definition of the titles of each column
            ws = wb.create_sheet("Temperature")
            ws.append(self.header)
            ws = wb.create_sheet("BarPressure")
            ws.append(self.header)
            ws = wb.create_sheet("Humidity")
            ws.append(self.header)

            #Delete the firstly created "Sheet" 
            ws = wb["Sheet"]
            wb.remove(ws)

            #Save the file
            wb.save("data.xlsx")
            #print(wb.sheetnames) #tabs check
    
       
    def csvR(self,data):
        with open("data.csv" , "a") as csv_file_obj:
            csv_writer = csv.writer(csv_file_obj, delimiter = ",")
            #csv_writer = csv.writer(csv_file_obj, delimiter = "\t")
            csv_writer.writerow(data)
            csv_file_obj.close()
        
    
    #saving data by row      
    def excelR(self, data,sheetname):
        wb = openpyxl.load_workbook("data.xlsx") #Workbook() to open a new workbook, but the existing gets deleted
        wb.save("data.xlsx")

        #loading the file again to save the data into the excel file 
        wb = openpyxl.load_workbook("data.xlsx")
        ws = wb[sheetname]
        ws.append(data)
        wb.save("data.xlsx")


######################################################


#Data sampling
class DataClassification(StoreData):
    #attributes - can be included into the obj params
    #N = 10   #100 measurements in order to meet the CI requirement - set at the main sketch
    #time_intervals = 1 - set at the main sketch
    #horizon = N*time_intervals/60

    #Attribute
    Sampling = []
    Sampling = []
    header   = ['Current_time','Lower_confidence_interval_value',
                  'Upper_confidence_interval_value','Mean_temperature_value' ,
                  'Median_temperature_value']
    
    #dictionary that will document all data
    Data_temp   = { header[0]: [], header[1]: [], header[2]: [], header[3]: [] ,
                  header[4]: []}
    Data_bpress = { header[0]: [], header[1]: [], header[2]: [], header[3]: [] ,
                  header[4]: []}
    Data_hum    = { header[0]: [], header[1]: [], header[2]: [], header[3]: [] ,
                  header[4]: []}
    att = 'csvf'
    
    def __init__(self, N ,  time_intervals, dictionary , att):    
        #attribute = 'csvf'  #defines the form of the file that we need to have csv / xls
        
        #Class parent inheritance using super dumber method using the header
        #attribute that has been allocated in this class
        super().__init__(self.header , att)
        self.Data_temp
        self.Data_bpress
        self.Data_hum
        self.N = N
        self.time_intervals = time_intervals
        self.dictionary = dictionary

        
    def csv_data_doc(self,instance_meas):
        if len(self.Sampling) == self.N:
            datapT = DataProcessing([row[0] for row in self.Sampling])
            datapP = DataProcessing([row[1] for row in self.Sampling])
            datapH = DataProcessing([row[2] for row in self.Sampling])
            #temporary list of the necessary processed data over N sampling
            now = datetime.now()
            current_time_p = now.strftime("%d/%m/%Y_%H:%M:%S")
            temp_val = [current_time_p , datapT.conf_Inter()[0] , datapT.conf_Inter()[1],
                        datapT.meanv() , datapT.medianv() , datapP.conf_Inter()[0] , datapP.conf_Inter()[1] , 
                        datapP.meanv() , datapP.medianv() , datapH.conf_Inter()[0] , datapH.conf_Inter()[1] , 
                        datapH.meanv() , datapH.medianv()]
            
            #Inheritance method from data class - to be saved in a csv file
            self.csvR(temp_val)

            self.Sampling = []

        self.Sampling.append(instance_meas)
        
        
    def temperature(self, instance_meas):
        if len(self.Sampling) == self.N:
            datap = DataProcessing(self.Sampling)
            #temporary list of the necessary processed data over N sampling
            now = datetime.now()
            current_time_p = now.strftime("%d/%m/%Y_%H:%M:%S")
            temp_val = [current_time_p,datap.conf_Inter()[0] , datap.conf_Inter()[1],
                        datap.meanv() , datap.medianv() ]
            #Inheritance from sheetdata class
            self.excelR(temp_val, "Temperature")
            #print(f' Temperature : {temp_val}') #measurements check

            if self.dictionary == True:
                for i in range(len(self.header)):
                    self.Data_temp[self.header[i]].append(temp_val[i])
            self.Sampling = []

        self.Sampling.append(instance_meas)
       

    def barpressure(self, instance_meas):  
        if len(self.Sampling) == self.N:
            datap = DataProcessing(self.Sampling)
            #temporary list of the necessary processed data over N sampling
            now = datetime.now()
            current_time_p = now.strftime("%d/%m/%Y_%H:%M:%S")
            temp_val = [current_time_p,datap.conf_Inter()[0] , datap.conf_Inter()[1],
                        datap.meanv() , datap.medianv() ]
            #Inheritance from sheetdata class
            self.excelR(temp_val, "BarPressure")
            #print(f'Barometric pressure : {temp_val}') #measurements check

            if self.dictionary == True:
                for i in range(len(self.header)):
                    self.Data_bpress[self.header[i]].append(temp_val[i])
            self.Sampling = []
            
        self.Sampling.append(instance_meas)
        
        
    def humidity(self, instance_meas):  
        if len(self.Sampling) == self.N:
            datap = DataProcessing(self.Sampling)
            #temporary list of the necessary processed data over N sampling
            now = datetime.now()
            current_time_p = now.strftime("%d/%m/%Y_%H:%M:%S")
            temp_val = [current_time_p,datap.conf_Inter()[0] , datap.conf_Inter()[1],
                        datap.meanv() , datap.medianv() ]
            #Inheritance from sheetdata class
            self.excelR(temp_val, "Humidity")
            #print(f' Humidity : {temp_val}') #measurements check

            if self.dictionary == True:
                for i in range(len(self.header)):
                    self.Data_hum[self.header[i]].append(temp_val[i])
            self.Sampling = []
            
        self.Sampling.append(instance_meas)


#####################################################


#Process data over horizon and output a single value after N samplings       
class DataProcessing():
    def __init__(self,N_samples):
        self.N_samples = N_samples #Will be going on every N data period

    #Processing and saving data using 3 methods mean, median, CI
    def meanv(self):
        return np.mean(self.N_samples)

    def medianv(self):
        return np.median(self.N_samples)
    
    def conf_Inter(self):
        avg=np.average(self.N_samples)
        std=np.std(self.N_samples)
        #t-student for 99% confidence interval gives upper and lower bound 
        min_val=avg-0.166*std*math.sqrt(100/99)
        max_val=avg+0.166*std*math.sqrt(100/99)
        return [min_val , max_val]
    
